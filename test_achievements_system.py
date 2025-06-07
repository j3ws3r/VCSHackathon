#!/usr/bin/env python3
"""
Comprehensive Test Suite for Achievements System
Tests both app/api/achievements.py and app/services/archievements_import.py
"""

import pytest
import os
import tempfile
import openpyxl
from io import BytesIO
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from unittest.mock import Mock, patch, MagicMock

# Import the modules to test
from app.main import app
from app.services.archievements_import import (
    AchievementImport, parse_excel, bulk_insert_achievements
)
from app.models.achievements import Achievement
from app.core.database import get_db, Base

# Test client
client = TestClient(app)

# Test database setup
SQLALCHEMY_DATABASE_URL = "sqlite:///./test_achievements.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def override_get_db():
    try:
        db = TestingSessionLocal()
        yield db
    finally:
        db.close()

app.dependency_overrides[get_db] = override_get_db

class TestAchievementImport:
    """Test the AchievementImport Pydantic model"""
    
    def test_valid_achievement_import(self):
        """Test creating a valid AchievementImport object"""
        data = {
            "title": "Daily Exercise",
            "description": "Complete 30 minutes of exercise",
            "frequency": "daily",
            "duration": 30,
            "points_value": 10
        }
        achievement = AchievementImport(**data)
        assert achievement.title == "Daily Exercise"
        assert achievement.description == "Complete 30 minutes of exercise"
        assert achievement.frequency == "daily"
        assert achievement.duration == 30
        assert achievement.points_value == 10

    def test_invalid_achievement_import_negative_duration(self):
        """Test validation fails for negative duration"""
        data = {
            "title": "Invalid Achievement",
            "description": "Test description",
            "frequency": "daily",
            "duration": -5,  # Invalid: negative
            "points_value": 10
        }
        with pytest.raises(Exception):  # Should raise validation error
            AchievementImport(**data)

    def test_invalid_achievement_import_negative_points(self):
        """Test validation fails for negative points"""
        data = {
            "title": "Invalid Achievement",
            "description": "Test description", 
            "frequency": "daily",
            "duration": 30,
            "points_value": -1  # Invalid: negative
        }
        with pytest.raises(Exception):  # Should raise validation error
            AchievementImport(**data)

    def test_achievement_import_max_title_length(self):
        """Test title length validation"""
        data = {
            "title": "A" * 256,  # Exceeds max length of 255
            "description": "Test description",
            "frequency": "daily", 
            "duration": 30,
            "points_value": 10
        }
        with pytest.raises(Exception):  # Should raise validation error
            AchievementImport(**data)

class TestExcelParsing:
    """Test Excel file parsing functionality"""

    def create_test_excel(self, data_rows):
        """Helper method to create test Excel files"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Header row
        headers = ["Title", "Description", "Frequency", "Duration", "Points"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Data rows
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        workbook.save(temp_file.name)
        temp_file.close()
        return temp_file.name

    def test_parse_valid_excel(self):
        """Test parsing a valid Excel file"""
        test_data = [
            ("Daily Exercise", "30 minutes of physical activity", "daily", 30, 10),
            ("Weekly Reading", "Read for 2 hours", "weekly", 120, 25),
            ("Monthly Goal", "Complete monthly project", "monthly", 30, 50)
        ]
        
        excel_file = self.create_test_excel(test_data)
        
        try:
            parsed_data = parse_excel(excel_file)
            
            assert len(parsed_data) == 3
            
            # Test first achievement
            assert parsed_data[0].title == "Daily Exercise"
            assert parsed_data[0].description == "30 minutes of physical activity"
            assert parsed_data[0].frequency == "daily"
            assert parsed_data[0].duration == 30
            assert parsed_data[0].points_value == 10
            
            # Test second achievement
            assert parsed_data[1].title == "Weekly Reading"
            assert parsed_data[1].duration == 120
            assert parsed_data[1].points_value == 25
            
        finally:
            os.unlink(excel_file)

    def test_parse_excel_with_invalid_rows(self):
        """Test parsing Excel with some invalid rows"""
        test_data = [
            ("Valid Achievement", "Good description", "daily", 30, 10),
            ("Invalid Duration", "Bad data", "daily", -5, 10),  # Invalid duration
            ("Invalid Points", "Bad points", "weekly", 60, -5),  # Invalid points
            ("Valid Achievement 2", "Another good one", "monthly", 45, 20)
        ]
        
        excel_file = self.create_test_excel(test_data)
        
        try:
            with patch('builtins.print') as mock_print:
                parsed_data = parse_excel(excel_file)
                
                # Should only parse valid rows (first and last)
                assert len(parsed_data) == 2
                assert parsed_data[0].title == "Valid Achievement"
                assert parsed_data[1].title == "Valid Achievement 2"
                
                # Should have printed error messages for invalid rows
                assert mock_print.call_count >= 2  # At least 2 error prints
                
        finally:
            os.unlink(excel_file)

    def test_parse_empty_excel(self):
        """Test parsing an empty Excel file"""
        excel_file = self.create_test_excel([])
        
        try:
            parsed_data = parse_excel(excel_file)
            assert len(parsed_data) == 0
            
        finally:
            os.unlink(excel_file)

class TestDatabaseOperations:
    """Test database operations for achievements"""

    def setup_method(self):
        """Setup test database before each test"""
        Base.metadata.create_all(bind=engine)

    def teardown_method(self):
        """Clean up test database after each test"""
        Base.metadata.drop_all(bind=engine)

    def test_bulk_insert_achievements_new_data(self):
        """Test inserting new achievements into database"""
        db = TestingSessionLocal()
        
        test_achievements = [
            AchievementImport(
                title="Test Achievement 1",
                description="First test achievement",
                frequency="daily",
                duration=30,
                points_value=10
            ),
            AchievementImport(
                title="Test Achievement 2", 
                description="Second test achievement",
                frequency="weekly",
                duration=60,
                points_value=25
            )
        ]
        
        try:
            result = bulk_insert_achievements(test_achievements, db)
            
            assert result["created"] == 2
            assert result["skipped"] == 0
            
            # Verify data was inserted
            all_achievements = db.query(Achievement).all()
            assert len(all_achievements) == 2
            
            achievement1 = db.query(Achievement).filter(Achievement.title == "Test Achievement 1").first()
            assert achievement1 is not None
            assert achievement1.description == "First test achievement"
            assert achievement1.points_value == 10
            
        finally:
            db.close()

    def test_bulk_insert_achievements_with_duplicates(self):
        """Test inserting achievements with some duplicates"""
        db = TestingSessionLocal()
        
        # First, insert an achievement
        existing_achievement = Achievement(
            title="Existing Achievement",
            description="Already exists",
            frequency="daily",
            duration=30,
            points_value=15
        )
        db.add(existing_achievement)
        db.commit()
        
        # Now try to insert new data including the duplicate
        test_achievements = [
            AchievementImport(
                title="Existing Achievement",  # Duplicate
                description="Duplicate description",
                frequency="daily",
                duration=30,  # Same duration = duplicate
                points_value=20
            ),
            AchievementImport(
                title="New Achievement",
                description="This is new",
                frequency="weekly", 
                duration=45,
                points_value=30
            )
        ]
        
        try:
            result = bulk_insert_achievements(test_achievements, db)
            
            assert result["created"] == 1  # Only the new one
            assert result["skipped"] == 1  # The duplicate
            
            # Verify total count
            all_achievements = db.query(Achievement).all()
            assert len(all_achievements) == 2  # Original + new one
            
        finally:
            db.close()

class TestAchievementsAPIEndpoints:
    """Test the FastAPI endpoints in achievements.py"""

    def setup_method(self):
        """Setup test database before each test"""
        Base.metadata.create_all(bind=engine)

    def teardown_method(self):
        """Clean up test database after each test"""
        Base.metadata.drop_all(bind=engine)

    def create_test_excel_file(self):
        """Create a test Excel file for upload testing"""
        test_data = [
            ("API Test Achievement", "Testing via API", "daily", 45, 15),
            ("Second API Achievement", "Another test", "weekly", 90, 35)
        ]
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        headers = ["Title", "Description", "Frequency", "Duration", "Points"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
            
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to BytesIO for upload
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    @patch('app.api.routes.get_current_user')
    def test_upload_excel_endpoint_success(self, mock_auth):
        """Test successful Excel file upload"""
        # Mock authentication
        mock_auth.return_value = {"id": 1, "email": "test@example.com", "role": "user"}
        
        excel_file = self.create_test_excel_file()
        
        response = client.post(
            "/api/v1/achievements/upload",
            files={"file": ("test_achievements.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        json_response = response.json()
        assert "message" in json_response
        assert "filename" in json_response
        assert "being processed" in json_response["message"]

    @patch('app.api.routes.get_current_user')
    def test_upload_invalid_file_type(self, mock_auth):
        """Test upload with invalid file type"""
        mock_auth.return_value = {"id": 1, "email": "test@example.com", "role": "user"}
        
        # Create a text file instead of Excel
        text_file = BytesIO(b"This is not an Excel file")
        
        response = client.post(
            "/api/v1/achievements/upload",
            files={"file": ("test.txt", text_file, "text/plain")}
        )
        
        assert response.status_code == 400
        assert "Invalid file type" in response.json()["detail"]

    @patch('app.api.routes.get_current_user')
    def test_preview_excel_endpoint_success(self, mock_auth):
        """Test successful Excel file preview"""
        mock_auth.return_value = {"id": 1, "email": "test@example.com", "role": "user"}
        
        excel_file = self.create_test_excel_file()
        
        response = client.post(
            "/api/v1/achievements/preview",
            files={"file": ("test_achievements.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        json_response = response.json()
        assert "preview" in json_response
        assert len(json_response["preview"]) <= 10  # Limited to 10 items
        
        # Check first item structure
        if json_response["preview"]:
            first_item = json_response["preview"][0]
            assert "title" in first_item
            assert "description" in first_item
            assert "frequency" in first_item
            assert "duration" in first_item
            assert "points_value" in first_item

    @patch('app.api.routes.get_current_user')
    def test_preview_invalid_file_type(self, mock_auth):
        """Test preview with invalid file type"""
        mock_auth.return_value = {"id": 1, "email": "test@example.com", "role": "user"}
        
        text_file = BytesIO(b"Not an Excel file")
        
        response = client.post(
            "/api/v1/achievements/preview",
            files={"file": ("test.txt", text_file, "text/plain")}
        )
        
        assert response.status_code == 400
        assert "Invalid file type" in response.json()["detail"]

    def test_upload_without_authentication(self):
        """Test upload endpoint without authentication"""
        excel_file = self.create_test_excel_file()
        
        response = client.post(
            "/api/v1/achievements/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Should return 401 or 403 (authentication required)
        assert response.status_code in [401, 403]

class TestIntegrationWorkflow:
    """Test complete workflow from upload to database"""

    def setup_method(self):
        """Setup test database before each test"""
        Base.metadata.create_all(bind=engine)

    def teardown_method(self):
        """Clean up test database after each test"""
        Base.metadata.drop_all(bind=engine)

    def test_complete_import_workflow(self):
        """Test the complete workflow from Excel parsing to database insertion"""
        # Create test Excel file
        test_data = [
            ("Workflow Test 1", "First workflow achievement", "daily", 20, 5),
            ("Workflow Test 2", "Second workflow achievement", "weekly", 60, 20),
            ("Duplicate Test", "This will be inserted", "monthly", 30, 15),
            ("Duplicate Test", "This will be skipped", "monthly", 30, 25)  # Same title+duration
        ]
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        headers = ["Title", "Description", "Frequency", "Duration", "Points"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
            
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        workbook.save(temp_file.name)
        temp_file.close()
        
        db = TestingSessionLocal()
        
        try:
            # Parse Excel
            parsed_data = parse_excel(temp_file.name)
            assert len(parsed_data) == 4  # All rows should parse successfully
            
            # Insert into database
            result = bulk_insert_achievements(parsed_data, db)
            assert result["created"] == 3  # 3 unique achievements
            assert result["skipped"] == 1   # 1 duplicate
            
            # Verify in database
            all_achievements = db.query(Achievement).all()
            assert len(all_achievements) == 3
            
            # Check specific records
            workflow1 = db.query(Achievement).filter(Achievement.title == "Workflow Test 1").first()
            assert workflow1 is not None
            assert workflow1.points_value == 5
            
            duplicate_count = db.query(Achievement).filter(Achievement.title == "Duplicate Test").count()
            assert duplicate_count == 1  # Only one should exist
            
        finally:
            db.close()
            os.unlink(temp_file.name)

def run_tests():
    """Run all tests with detailed output"""
    print("🧪 Starting Achievements System Test Suite")
    print("=" * 60)
    
    # Run tests with pytest
    pytest_args = [
        __file__,
        "-v",  # Verbose output
        "--tb=short",  # Short traceback format
        "--color=yes"  # Colored output
    ]
    
    exit_code = pytest.main(pytest_args)
    
    print("\n" + "=" * 60)
    if exit_code == 0:
        print("🎉 All tests passed successfully!")
    else:
        print("❌ Some tests failed. Check output above.")
    
    return exit_code

if __name__ == "__main__":
    run_tests() 